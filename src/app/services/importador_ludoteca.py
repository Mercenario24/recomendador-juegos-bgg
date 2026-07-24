import csv
import re
import unicodedata
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook


COLUMNAS_ID = [
    "objectid",
    "idbgg",
    "bggid",
    "gameid",
    "id"
]

COLUMNAS_NOMBRE = [
    "objectname",
    "name",
    "nombre",
    "game"
]

COLUMNAS_PROPIO = [
    "own",
    "owned",
    "propio"
]

COLUMNAS_VALORACION = [
    "rating",
    "userrating",
    "valoracion",
    "puntuacion"
]

COLUMNAS_PARTIDAS = [
    "numplays",
    "plays",
    "partidas"
]


def normalizar_nombre_columna(nombre):
    texto = str(nombre or "").strip().lower()

    texto = unicodedata.normalize(
        "NFKD",
        texto
    ).encode(
        "ascii",
        "ignore"
    ).decode("ascii")

    return re.sub(
        r"[^a-z0-9]",
        "",
        texto
    )


def buscar_columna(columnas, nombres_posibles):
    columnas_normalizadas = {
        normalizar_nombre_columna(columna): columna
        for columna in columnas
    }

    for nombre in nombres_posibles:
        nombre_normalizado = normalizar_nombre_columna(
            nombre
        )

        if nombre_normalizado in columnas_normalizadas:
            return columnas_normalizadas[nombre_normalizado]

    return None


def convertir_entero(valor, defecto=None):
    if valor is None:
        return defecto

    texto = str(valor).strip()

    if not texto:
        return defecto

    try:
        return int(float(texto.replace(",", ".")))
    except ValueError:
        return defecto


def convertir_decimal(valor):
    if valor is None:
        return None

    texto = str(valor).strip()

    if not texto:
        return None

    try:
        numero = float(texto.replace(",", "."))

        if numero < 0:
            return None

        return numero

    except ValueError:
        return None


def es_verdadero(valor):
    if valor is None:
        return False

    texto = str(valor).strip().lower()

    return texto in {
        "1",
        "1.0",
        "true",
        "yes",
        "y",
        "si",
        "sí",
        "s",
        "x"
    }


def decodificar_csv(contenido):
    codificaciones = [
        "utf-8-sig",
        "utf-8",
        "cp1252",
        "latin-1"
    ]

    for codificacion in codificaciones:
        try:
            return contenido.decode(codificacion)
        except UnicodeDecodeError:
            continue

    raise ValueError(
        "No se ha podido reconocer la codificación del CSV."
    )


def leer_filas_csv(contenido):
    texto = decodificar_csv(contenido)

    try:
        dialecto = csv.Sniffer().sniff(
            texto[:5000],
            delimiters=",;\t"
        )

        lector = csv.DictReader(
            StringIO(texto),
            dialect=dialecto
        )

    except csv.Error:
        lector = csv.DictReader(
            StringIO(texto),
            delimiter=","
        )

    return [
        dict(fila)
        for fila in lector
    ]


def leer_filas_xlsx(contenido):
    libro = load_workbook(
        filename=BytesIO(contenido),
        read_only=True,
        data_only=True
    )

    hoja = libro.active
    iterador_filas = hoja.iter_rows(values_only=True)

    try:
        cabeceras = next(iterador_filas)
    except StopIteration:
        libro.close()
        return []

    cabeceras = [
        str(cabecera).strip()
        if cabecera is not None
        else ""
        for cabecera in cabeceras
    ]

    filas = []

    for valores in iterador_filas:
        fila = {
            cabeceras[indice]: valor
            for indice, valor in enumerate(valores)
            if indice < len(cabeceras)
            and cabeceras[indice]
        }

        if any(
            valor is not None
            and str(valor).strip()
            for valor in fila.values()
        ):
            filas.append(fila)

    libro.close()

    return filas


def procesar_filas(filas):
    if not filas:
        raise ValueError(
            "El archivo no contiene datos."
        )

    columnas = list(filas[0].keys())

    columna_id = buscar_columna(
        columnas,
        COLUMNAS_ID
    )

    if columna_id is None:
        raise ValueError(
            "No se ha encontrado la columna con el ID de BGG. "
            "Se esperaba una columna como 'objectid' o 'id_bgg'."
        )

    columna_nombre = buscar_columna(
        columnas,
        COLUMNAS_NOMBRE
    )

    columna_propio = buscar_columna(
        columnas,
        COLUMNAS_PROPIO
    )

    columna_valoracion = buscar_columna(
        columnas,
        COLUMNAS_VALORACION
    )

    columna_partidas = buscar_columna(
        columnas,
        COLUMNAS_PARTIDAS
    )

    juegos_por_id = {}

    for fila in filas:
        if (
            columna_propio is not None
            and not es_verdadero(fila.get(columna_propio))
        ):
            continue

        id_bgg = convertir_entero(
            fila.get(columna_id)
        )

        if id_bgg is None or id_bgg <= 0:
            continue

        nombre = None

        if columna_nombre:
            nombre = str(
                fila.get(columna_nombre) or ""
            ).strip()

        valoracion = None

        if columna_valoracion:
            valoracion = convertir_decimal(
                fila.get(columna_valoracion)
            )

        num_partidas = 0

        if columna_partidas:
            num_partidas = convertir_entero(
                fila.get(columna_partidas),
                defecto=0
            )

        juegos_por_id[id_bgg] = {
            "id_bgg": id_bgg,
            "nombre_archivo": nombre,
            "valoracion_usuario": valoracion,
            "num_partidas": max(0, num_partidas)
        }

    juegos = list(juegos_por_id.values())

    if not juegos:
        raise ValueError(
            "No se han encontrado juegos propios en el archivo."
        )

    return juegos


def leer_archivo_ludoteca(contenido, nombre_archivo):
    extension = Path(
        nombre_archivo
    ).suffix.lower()

    if extension == ".csv":
        filas = leer_filas_csv(contenido)

    elif extension == ".xlsx":
        filas = leer_filas_xlsx(contenido)

    else:
        raise ValueError(
            "El archivo debe tener extensión .csv o .xlsx."
        )

    return procesar_filas(filas)